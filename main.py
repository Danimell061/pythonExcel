import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

try:
    planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
    pagina_fechamento = planilha_fechamento['Sheet1']
    nomes_existentes = {linha[0].value for linha in pagina_fechamento.iter_rows(min_row=2)}
except FileNotFoundError:
    nomes_existentes = set()

# 1 - Entrar na planilha e extrair o cpf do cliente
planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']
driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')

for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha
    if nome in nomes_existentes:
        continue
    # 2 - Entrar no site e usar o cpf da planilha pra pesquisar o status do pagamento daquele cliente
    sleep(5.0)
    campo_pesquisa = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
    sleep(1)
    campo_pesquisa.clear()
    campo_pesquisa.send_keys(cpf)
    sleep(1)
    #3 - Verificar se está "em dia" ou "atrasado"
    botao_pesquisar = driver.find_element(By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(1)
    botao_pesquisar.click()
    sleep(4)
    status = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
    if status.text == 'em dia':
        # 4 - Se estiver "em dia", pegar a data do pagamento e o metodo de pagamento colocar status pago
        data_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentDate']")
        metodo_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentMethod']")

        data_pagamento_limpo = data_pagamento.text.split()[3]
        metodo_pagamento_limpo = metodo_pagamento.text.split()[3]

        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']

        pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia', data_pagamento_limpo, metodo_pagamento_limpo])

        planilha_fechamento.save('planilha fechamento.xlsx')
        nomes_existentes.add(nome)
    else:
        # 5 - Se estiver "atrasado", colocar status atrasado
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']

        pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])
        planilha_fechamento.save('planilha fechamento.xlsx')
        nomes_existentes.add(nome)